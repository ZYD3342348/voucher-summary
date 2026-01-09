#!/usr/bin/env python3
"""
简易报表引擎雏形：
- 从工作表长表化（名称清洗、自动金额列、收入类型首字母）
- 过滤收入类型（默认 H），生成名称×项目透视
- 生成“调整表”（不计税/5%/6% 分桶，供人工覆盖）
- 汇总调整表，输出税分拆摘要

用法示例：
    python3 report_engine.py -i "2025年8月总台.xlsx" -w "工作表" -o "2025年8月报表.xlsx" -t H
"""

import argparse
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

# 引入现有工具
BASE = Path(__file__).resolve().parent.parent
sys.path.append(str(BASE))
from scripts.normalize_work import normalize_work  # type: ignore


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(index="名称", columns="项目", values="金额", aggfunc="sum", fill_value=0)
    pivot["总计"] = pivot.sum(axis=1)
    pivot.loc["合计"] = pivot.sum()
    return pivot.round(2)


def build_adjust_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    将名称×项目×金额展开，预填分桶：
    - 不计税收入：0
    - 计税收入-5%：0
    - 计税收入-6%：金额（默认全计税6%，便于人工改）
    """
    g = df.groupby(["名称", "项目"], as_index=False)["金额"].sum()
    adj = g.copy()
    adj["不计税收入"] = 0.0
    adj["计税收入-5%"] = 0.0
    adj["计税收入-6%"] = adj["金额"].astype(float)
    adj["备注"] = ""
    return adj


def summarize_tax(adj: pd.DataFrame) -> pd.DataFrame:
    """汇总调整表为简易税分拆摘要。"""
    rows = []
    notax = adj["不计税收入"].sum()
    tax5 = adj["计税收入-5%"].sum()
    tax6 = adj["计税收入-6%"].sum()

    def net_tax(amount, rate):
        if amount <= 0:
            return 0.0, 0.0
        net = amount / (1 + rate)
        tax = amount - net
        return net, tax

    net5, taxamt5 = net_tax(tax5, 0.05)
    net6, taxamt6 = net_tax(tax6, 0.06)

    rows.append(["不计税收入", round(notax, 2), "", ""])
    rows.append(["计税收入-5%", round(tax5, 2), round(net5, 2), round(taxamt5, 2)])
    rows.append(["计税收入-6%", round(tax6, 2), round(net6, 2), round(taxamt6, 2)])
    rows.append(["合计", round(notax + tax5 + tax6, 2), round(net5 + net6, 2), round(taxamt5 + taxamt6, 2)])

    return pd.DataFrame(rows, columns=["类别", "含税收入", "不含税收入", "税额"]).round(2)


def write_df(ws, df: pd.DataFrame, start_row=1, start_col=1):
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=col)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=val)


def main():
    ap = argparse.ArgumentParser(description="简易报表生成器（透视+调整表+税分拆）")
    ap.add_argument("-i", "--input", required=True, help="输入 Excel 文件（含工作表）")
    ap.add_argument("-w", "--work-sheet", default="工作表", help="工作表名，默认 工作表")
    ap.add_argument("-o", "--output", help="输出 Excel 文件")
    ap.add_argument("-t", "--income-type", default="H", help="收入类型过滤，默认 H")
    args = ap.parse_args()

    inp = Path(args.input)
    out = Path(args.output) if args.output else inp.with_name(f"{inp.stem}_报表.xlsx")

    # 读取并过滤
    df = normalize_work(inp, args.work_sheet)
    df = df[df["收入类型"].astype(str).str.upper() == args.income_type.upper()]

    # 透视
    pivot = build_pivot(df)
    # 调整表
    adj = build_adjust_table(df)
    # 摘要
    summary = summarize_tax(adj)

    wb = Workbook()
    ws_pivot = wb.active
    ws_pivot.title = "透视"
    write_df(ws_pivot, pivot.reset_index().rename(columns={"index": "名称"}))

    ws_adj = wb.create_sheet("调整表")
    write_df(ws_adj, adj)

    ws_sum = wb.create_sheet("税分拆摘要")
    write_df(ws_sum, summary)

    wb.save(out)
    print(f"[info] 已生成 {out}；透视行数 {len(pivot)}，调整行数 {len(adj)}")


if __name__ == "__main__":
    main()
