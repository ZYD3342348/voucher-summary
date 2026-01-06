#!/usr/bin/env python3
"""
总台财务数据汇总生成器
=========================
生成包含透视表、凭证说明、调整S计算、验证链路的汇总Excel文件。

特点：
- 公式引用动态生成：所有引用通过行号变量计算，避免硬编码（如C20误指向B24）。
- 单表输出：配置区、透视表、凭证说明、验证都在同一工作表中。
- 可复用：更换数据源后仍能保持公式链正确。

用法：
    python generate_summary.py

需要的输入：
    - INPUT_FILE: 源数据Excel（含明细与“收入类型表”）
    - SHEET_NAME: 要读取的工作表名（默认“收入类型表”）
    - TRANSFER_AMOUNT: 转账金额（可来自“总数”表）
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ============= 配置区 =============
INPUT_FILE = '12月总台.XLS.xlsx'           # 输入数据文件
OUTPUT_FILE = '12月汇总输出.xlsx'           # 输出文件
TRANSFER_AMOUNT = 600240                   # 转账金额（从总数表获取）
SHEET_NAME = '收入类型表'                  # 工作表名称
# ==================================


def create_workbook():
    """创建工作簿并设置样式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    # 设置列宽
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=11)
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    return wb, thin_border, header_font, green_fill


def load_income_data(input_file, sheet_name):
    """
    读取收入类型表的明细数据（透视表上方的明细区域）。
    自动识别透视表起始行：第一列等于“收入类型”的行视为透视表表头。
    """
    raw = pd.read_excel(input_file, sheet_name=sheet_name, header=None)

    pivot_start = None
    for idx, val in raw[0].items():
        if str(val).strip() == '收入类型':
            pivot_start = idx
            break

    header = raw.iloc[0]
    data_end = pivot_start if pivot_start is not None else len(raw)
    data = raw.iloc[1:data_end].copy()  # 跳过表头行
    data.columns = header

    # 只保留核心列
    data = data[['项目', '名称', '收入类型', '金额']]
    data['金额'] = pd.to_numeric(data['金额'], errors='coerce').fillna(0)
    return data


def build_pivot(df, income_types):
    """根据明细数据生成透视表，并补齐缺失的收入类型行。"""
    pivot = df.pivot_table(
        index='收入类型',
        columns='项目',
        values='金额',
        aggfunc='sum',
        fill_value=0
    )

    for itype in income_types:
        if itype not in pivot.index:
            pivot.loc[itype] = 0

    # 保持行顺序
    pivot = pivot.reindex(income_types)
    return pivot


def coord(row, col):
    """行列转单元格坐标（例如 row=7, col=2 -> B7）"""
    return f"{get_column_letter(col)}{row}"


def generate_summary(input_file, output_file, transfer_amount, sheet_name='收入类型表'):
    """生成汇总表，所有公式引用动态确定。"""

    print(f"读取数据: {input_file}")
    print(f"工作表: {sheet_name}")
    print(f"转账金额: {transfer_amount:,}")
    print("-" * 50)

    income_types = ['H', 'L', 'R', 'S', 'T', 'Z']

    # 读明细并生成透视
    detail_df = load_income_data(input_file, sheet_name)
    pivot = build_pivot(detail_df, income_types)

    # 备好每个类型的房费和总计
    type_room_fee = {t: float(pivot.loc[t].get('房费', 0)) for t in income_types}
    type_total = {t: float(pivot.loc[t].sum()) for t in income_types}
    total_room_value = sum(type_room_fee.values())

    wb, thin_border, header_font, green_fill = create_workbook()
    ws = wb.active

    # ========== 配置区 ==========
    ws['A1'] = '转账金额'
    ws['B1'] = transfer_amount
    ws['A2'] = '数据文件'
    ws['B2'] = input_file

    # ========== 透视表标题 ==========
    pivot_header_row = 6
    headers = ['收入类型', '房费', '其他项目合计', '总计']
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=pivot_header_row, column=idx, value=title)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ========== 透视表数据 ==========
    pivot_rows = {}
    current_row = pivot_header_row
    for itype in income_types:
        current_row += 1
        pivot_rows[itype] = current_row
        ws.cell(row=current_row, column=1, value=itype)

        room_fee = type_room_fee.get(itype, 0)
        total_val = type_total.get(itype, room_fee)

        ws.cell(row=current_row, column=2, value=room_fee)
        # 其他项目合计使用公式保持链路：总计 - 房费
        ws.cell(row=current_row, column=3, value=f"=D{current_row}-B{current_row}")
        ws.cell(row=current_row, column=4, value=total_val)

    # 房费总计行（公式汇总房费列）
    room_total_row = current_row + 1
    ws.cell(row=room_total_row, column=1, value='房费总计')
    ws.cell(row=room_total_row, column=2, value=f"=SUM(B{pivot_header_row+1}:B{room_total_row-1})")
    ws.cell(row=room_total_row, column=3, value=f"=B{room_total_row}")
    ws.cell(row=room_total_row, column=4, value=f"=B{room_total_row}")

    # 房费剩余行
    room_remaining_row = room_total_row + 1
    ws.cell(row=room_remaining_row, column=1, value='房费剩余')
    transfer_ref = '$B$1'  # 配置区转账金额引用
    ws.cell(row=room_remaining_row, column=2, value=f"=B{room_total_row}-{transfer_ref}")

    # 关键引用坐标
    room_total_cell = coord(room_total_row, 2)
    room_remaining_cell = coord(room_remaining_row, 2)
    h_fee_cell = coord(pivot_rows['H'], 2)
    l_fee_cell = coord(pivot_rows['L'], 2)
    t_fee_cell = coord(pivot_rows['T'], 2)

    # ========== 凭证说明 ==========
    voucher_title_row = room_remaining_row + 2
    ws.cell(row=voucher_title_row, column=1, value='凭证说明').font = header_font

    voucher_header_row = voucher_title_row + 1
    voucher_headers = ['收入类型', '凭证摘要', '金额', '说明']
    for idx, title in enumerate(voucher_headers, start=1):
        cell = ws.cell(row=voucher_header_row, column=idx, value=title)
        cell.font = header_font
        cell.border = thin_border

    current_row = voucher_header_row
    voucher_rows = {}

    def add_voucher_row(itype, desc, amount_formula, note):
        nonlocal current_row
        current_row += 1
        voucher_rows[itype] = current_row
        ws.cell(row=current_row, column=1, value=itype)
        ws.cell(row=current_row, column=2, value=desc)
        ws.cell(row=current_row, column=3, value=amount_formula)
        ws.cell(row=current_row, column=4, value=note)

    add_voucher_row('H', '总台收入-住宿费（会议类）', f'={h_fee_cell}', '会议/内部成本')
    add_voucher_row('L', '总台收入-住宿费', f'={l_fee_cell}', '老干部/财政')
    add_voucher_row('T', '总台收入-住宿费', f'={t_fee_cell}', '其他事业单位')

    # 调整S = 房费总计 - 转账 - H房费 - L房费 - T房费
    adjust_formula = f'={room_total_cell}-{transfer_ref}-{h_fee_cell}-{l_fee_cell}-{t_fee_cell}'
    add_voucher_row('调整S', '总台收入-住宿费', adjust_formula, '扣除转账和H/L/T后的剩余金额')

    # 凭证合计
    voucher_total_row = current_row + 1
    first_voucher_row = voucher_header_row + 1
    ws.cell(row=voucher_total_row, column=1, value='凭证合计')
    ws.cell(row=voucher_total_row, column=3, value=f"=SUM(C{first_voucher_row}:C{current_row})")

    voucher_total_cell = coord(voucher_total_row, 3)

    # ========== 验证 ==========
    validation_title_row = voucher_total_row + 2
    ws.cell(row=validation_title_row, column=1, value='验证').font = header_font

    ws.cell(row=validation_title_row + 1, column=1, value='房费剩余')
    ws.cell(row=validation_title_row + 1, column=2, value=f'={room_remaining_cell}')

    ws.cell(row=validation_title_row + 2, column=1, value='凭证合计')
    ws.cell(row=validation_title_row + 2, column=2, value=f'={voucher_total_cell}')

    result_cell = ws.cell(
        row=validation_title_row + 3,
        column=2,
        value=f'=IF({room_remaining_cell}={voucher_total_cell},"✓ 正确","✗ 错误")'
    )
    ws.cell(row=validation_title_row + 3, column=1, value='验证结果')
    result_cell.fill = green_fill

    # 输出统计信息
    print("-" * 50)
    print("数据统计:")
    print(f"  明细记录数: {len(detail_df):,}")
    print(f"  房费总计: {total_room_value:,.2f}")
    print(f"  转账金额: {transfer_amount:,}")
    print(f"  房费剩余: {total_room_value - transfer_amount:,.2f}")
    print("-" * 50)

    wb.save(output_file)
    print(f"✓ 已生成: {output_file}")

    return pivot_rows


if __name__ == '__main__':
    generate_summary(INPUT_FILE, OUTPUT_FILE, TRANSFER_AMOUNT, SHEET_NAME)
